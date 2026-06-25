#!/opt/homebrew/anaconda3/envs/novem12/bin/python
# -*- coding: utf-8 -*-
"""订单-停车费/违章匹配 桌面工具 (PySide6)

用法:
    /opt/homebrew/anaconda3/envs/novem12/bin/python order_match_gui.py
"""

from __future__ import annotations

import re
import sys
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from PySide6.QtCore import QThread, Signal
from PySide6.QtGui import QFont
from PySide6.QtWidgets import (
    QApplication,
    QFileDialog,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QVBoxLayout,
    QWidget,
)

# ============== 核心匹配逻辑(从 match_orders.py 复用) ==============


def norm_plate(s: str) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", "", str(s)).upper()


def to_dt(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v
    if isinstance(v, (int, float)):
        return datetime(1899, 12, 30) + timedelta(days=float(v))
    s = str(v).strip()
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y-%m-%d",
        "%Y/%m/%d",
    ):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def find_col_index(headers: list[str], col_name: str, sheet_name: str) -> int:
    """在表头中精确查定位所需字段的列索引（过滤掉首尾空格）"""
    cleaned_name = str(col_name).strip()
    for idx, h in enumerate(headers):
        if h is not None and str(h).strip() == cleaned_name:
            return idx
    raise ValueError(f"工作表【{sheet_name}】中未找到必须的列头：'{col_name}'")


def load_orders(path: str) -> tuple[dict[str, list[dict]], list[str]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    if "订单数据" not in wb.sheetnames:
        raise ValueError("源文件中缺少必填的【订单数据】工作表")
    ws = wb["订单数据"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("工作表【订单数据】为空")

    headers = [str(h).strip() for h in rows[0] if h is not None]

    # 精确匹配三大核心列
    plate_idx = find_col_index(headers, "最新车牌号", "订单数据")
    pick_idx = find_col_index(headers, "实际取车时间", "订单数据")
    ret_idx = find_col_index(headers, "实际还车时间", "订单数据")

    by_plate: dict[str, list[dict]] = {}
    for row in rows[1:]:
        if not row or all(v is None for v in row):
            continue
        plate = norm_plate(row[plate_idx] if plate_idx < len(row) else None)
        pick = to_dt(row[pick_idx] if pick_idx < len(row) else None)
        ret = to_dt(row[ret_idx] if ret_idx < len(row) else None)
        if not plate or not pick or not ret:
            continue

        # 动态保存一整行的所有字段
        row_dict = {}
        for idx, col_name in enumerate(headers):
            val = row[idx] if idx < len(row) else None
            # 存入解析后的 datetime 以便后续比较
            if idx == pick_idx:
                row_dict[col_name] = pick
            elif idx == ret_idx:
                row_dict[col_name] = ret
            else:
                row_dict[col_name] = val

        by_plate.setdefault(plate, []).append(row_dict)

    for v in by_plate.values():
        pick_col_name = headers[pick_idx]
        v.sort(key=lambda x: x[pick_col_name])

    return by_plate, headers


def process_sheet(
    src_ws, out_wb, orders_by_plate, order_headers, src_sheet_name, progress_cb=None
):
    rows = list(src_ws.iter_rows(values_only=True))
    if not rows:
        return 0, 0
    headers = [str(h).strip() for h in rows[0] if h is not None]
    data_rows = rows[1:]

    # 定位本表核心列
    plate_col = find_col_index(headers, "车牌号", src_sheet_name)

    if src_sheet_name == "停车费":
        in_col = find_col_index(headers, "入场时间", src_sheet_name)
        out_col = find_col_index(headers, "出场时间", src_sheet_name)
    else:  # 违章
        in_col = None
        out_col = None
        event_col = find_col_index(headers, "违法时间", src_sheet_name)

    # 订单表的关键匹配字段名
    order_pick_col = "实际取车时间"
    order_ret_col = "实际还车时间"

    # 将订单数据的所有列合并到当前匹配数据行的后面
    new_headers = list(headers) + list(order_headers)

    out_ws = out_wb.create_sheet(src_ws.title)
    out_ws.append(new_headers)

    # 复制原始列宽
    src_widths = {
        col: dim.width for col, dim in src_ws.column_dimensions.items() if dim.width
    }
    for col_letter, w in src_widths.items():
        out_ws.column_dimensions[col_letter].width = w

    # 订单列的列宽默认为 22
    for idx in range(len(headers) + 1, len(new_headers) + 1):
        col_letter = openpyxl.utils.get_column_letter(idx)
        out_ws.column_dimensions[col_letter].width = 22

    total = len(data_rows)
    matched = unmatched = 0

    for i, r in enumerate(data_rows):
        if r is None or all(v is None for v in r):
            continue
        plate = norm_plate(r[plate_col] if plate_col < len(r) else None)
        candidates = orders_by_plate.get(plate, [])

        if src_sheet_name == "停车费":
            t_in = to_dt(r[in_col] if in_col < len(r) else None)
            t_out = to_dt(r[out_col] if out_col < len(r) else None)
            # 停车费区间 与 订单 [实际取车时间, 实际还车时间] 有交集 → 匹配
            hits = (
                [
                    o
                    for o in candidates
                    if t_in < o[order_ret_col] and t_out > o[order_pick_col]
                ]
                if t_in and t_out
                else []
            )
        else:  # 违章
            t_event = to_dt(r[event_col] if event_col < len(r) else None)
            # 违章时间 落在 订单 [实际取车时间, 实际还车时间] 内 → 匹配
            hits = (
                [
                    o
                    for o in candidates
                    if o[order_pick_col] <= t_event <= o[order_ret_col]
                ]
                if t_event
                else []
            )

        if not hits:
            unmatched += 1
            new_row = list(r) + [None] * len(order_headers)
            while len(new_row) < len(new_headers):
                new_row.append(None)
            out_ws.append(new_row[: len(new_headers)])
        else:
            for o in hits:
                matched += 1
                new_row = list(r)
                while len(new_row) < len(headers):
                    new_row.append(None)
                # 动态追加全部订单列
                for oh in order_headers:
                    new_row.append(o.get(oh))
                out_ws.append(new_row[: len(new_headers)])

        if progress_cb and total:
            pct = (i + 1) / total
            progress_cb(pct, src_sheet_name, i + 1, total)

    return matched, unmatched


def run_match(src_path: str, dst_path: str, progress_cb=None):
    """主入口:读源文件 -> 匹配 -> 写新文件。progress_cb(pct, sheet, i, total)"""
    wb = openpyxl.load_workbook(src_path, data_only=True)
    orders_by_plate, order_headers = load_orders(src_path)
    total_orders = sum(len(v) for v in orders_by_plate.values())

    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)

    # 订单数据原样
    src_ws = wb["订单数据"]
    dst_ws = out_wb.create_sheet("订单数据")
    for row in src_ws.iter_rows(values_only=True):
        dst_ws.append(row)
    if progress_cb:
        progress_cb(0.0, "订单数据", 0, 1)

    def _half_cb(pct, sheet, i, total):
        if progress_cb is None:
            return
        if sheet == "停车费":
            progress_cb(pct * 0.5, sheet, i, total)
        else:
            progress_cb(0.5 + pct * 0.5, sheet, i, total)

    park_matched = park_unmatched = 0
    if "停车费" in wb.sheetnames:
        park_matched, park_unmatched = process_sheet(
            wb["停车费"],
            out_wb,
            orders_by_plate,
            order_headers,
            "停车费",
            progress_cb=_half_cb,
        )

    vio_matched = vio_unmatched = 0
    if "违章" in wb.sheetnames:
        vio_matched, vio_unmatched = process_sheet(
            wb["违章"],
            out_wb,
            orders_by_plate,
            order_headers,
            "违章",
            progress_cb=_half_cb,
        )

    out_wb.save(dst_path)

    return {
        "plates": len(orders_by_plate),
        "orders": total_orders,
        "park_rows": park_matched,
        "park_unmatched": park_unmatched,
        "vio_rows": vio_matched,
        "vio_unmatched": vio_unmatched,
    }


# ============== GUI ==============


class MatchWorker(QThread):
    """后台线程跑匹配,避免卡 UI"""

    progress = Signal(float, str, int, int)
    done = Signal(dict)
    failed = Signal(str)

    def __init__(self, src: str, dst: str):
        super().__init__()
        self.src = src
        self.dst = dst

    def run(self):
        try:
            result = run_match(self.src, self.dst, progress_cb=self._emit)
            self.done.emit(result)
        except Exception as e:
            import traceback

            self.failed.emit(f"{e}\n\n{traceback.format_exc()}")

    def _emit(self, pct, sheet, i, total):
        self.progress.emit(pct, sheet, i, total)


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("订单-停车费/违章 匹配工具")
        self.resize(720, 320)

        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)
        layout.setSpacing(12)
        layout.setContentsMargins(20, 20, 20, 20)

        title = QLabel("📋 订单-停车费/违章 匹配工具")
        title_font = QFont()
        title_font.setPointSize(16)
        title_font.setBold(True)
        title.setFont(title_font)
        layout.addWidget(title)

        hint = QLabel(
            "说明:\n"
            "  • 停车费区间 与 订单 [取车, 还车] 有交集 → 匹配\n"
            "  • 违章时间 落在 订单 [取车, 还车] 内 → 匹配\n"
            "  • 源文件需含 sheet: 订单数据 / 停车费 / 违章"
        )
        hint.setStyleSheet("color: #666;")
        layout.addWidget(hint)

        # 源文件
        src_row = QHBoxLayout()
        src_row.addWidget(QLabel("源文件:"))
        self.src_edit = QLineEdit()
        self.src_edit.setPlaceholderText(
            "选择含 订单数据/停车费/违章 三个 sheet 的 xlsx"
        )
        src_row.addWidget(self.src_edit, 1)
        self.src_btn = QPushButton("选择…")
        self.src_btn.clicked.connect(self.choose_src)
        src_row.addWidget(self.src_btn)
        layout.addLayout(src_row)

        # 输出文件
        dst_row = QHBoxLayout()
        dst_row.addWidget(QLabel("输出文件:"))
        self.dst_edit = QLineEdit()
        self.dst_edit.setPlaceholderText(
            "匹配结果输出位置(默认源文件名加 _匹配后 后缀)"
        )
        dst_row.addWidget(self.dst_edit, 1)
        self.dst_btn = QPushButton("选择…")
        self.dst_btn.clicked.connect(self.choose_dst)
        dst_row.addWidget(self.dst_btn)
        layout.addLayout(dst_row)

        # 进度条
        self.progress = QProgressBar()
        self.progress.setRange(0, 100)
        self.progress.setValue(0)
        self.progress.setFormat("%p%  (%v/100)")
        layout.addWidget(self.progress)

        self.status_label = QLabel("就绪")
        self.status_label.setStyleSheet("color: #444;")
        layout.addWidget(self.status_label)

        # 按钮
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        self.match_btn = QPushButton("开始匹配")
        self.match_btn.setMinimumHeight(38)
        self.match_btn.setMinimumWidth(140)
        self.match_btn.setStyleSheet(
            "QPushButton{background:#1976d2;color:white;font-weight:bold;border-radius:4px;}"
            "QPushButton:hover{background:#1565c0;}"
            "QPushButton:disabled{background:#90a4ae;}"
        )
        self.match_btn.clicked.connect(self.start_match)
        btn_row.addWidget(self.match_btn)
        layout.addLayout(btn_row)

        self.worker: MatchWorker | None = None
        self.setAcceptDrops(True)

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            urls = event.mimeData().urls()
            if urls and urls[0].toLocalFile().lower().endswith(".xlsx"):
                event.acceptProposedAction()

    def dropEvent(self, event):
        urls = event.mimeData().urls()
        if urls:
            path = urls[0].toLocalFile()
            if path:
                self.src_edit.setText(path)
                p = Path(path)
                self.dst_edit.setText(str(p.with_name(p.stem + "_匹配后.xlsx")))

    def choose_src(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "选择源文件", "", "Excel 文件 (*.xlsx);;所有文件 (*)"
        )
        if path:
            self.src_edit.setText(path)
            if not self.dst_edit.text():
                p = Path(path)
                self.dst_edit.setText(str(p.with_name(p.stem + "_匹配后.xlsx")))

    def choose_dst(self):
        path, _ = QFileDialog.getSaveFileName(
            self,
            "选择输出位置",
            self.dst_edit.text() or "匹配结果.xlsx",
            "Excel 文件 (*.xlsx)",
        )
        if path:
            if not path.lower().endswith(".xlsx"):
                path += ".xlsx"
            self.dst_edit.setText(path)

    def start_match(self):
        src = self.src_edit.text().strip()
        dst = self.dst_edit.text().strip()
        if not src:
            QMessageBox.warning(self, "提示", "请先选择源文件")
            return
        if not Path(src).exists():
            QMessageBox.warning(self, "提示", f"源文件不存在:\n{src}")
            return
        if not dst:
            p = Path(src)
            dst = str(p.with_name(p.stem + "_匹配后.xlsx"))
            self.dst_edit.setText(dst)

        self.match_btn.setEnabled(False)
        self.src_btn.setEnabled(False)
        self.dst_btn.setEnabled(False)
        self.progress.setValue(0)
        self.status_label.setText("开始匹配…")

        self.worker = MatchWorker(src, dst)
        self.worker.progress.connect(self.on_progress)
        self.worker.done.connect(self.on_done)
        self.worker.failed.connect(self.on_failed)
        self.worker.start()

    def on_progress(self, pct: float, sheet: str, i: int, total: int):
        self.progress.setValue(int(pct * 100))
        self.status_label.setText(f"正在处理 [{sheet}] — {i}/{total}")

    def on_done(self, result: dict):
        self.progress.setValue(100)
        self.status_label.setText("完成 ✅")
        self.match_btn.setEnabled(True)
        self.src_btn.setEnabled(True)
        self.dst_btn.setEnabled(True)
        msg = (
            f"匹配完成!\n\n"
            f"  • 订单数(有效): {result['orders']} 单 ({result['plates']} 个车牌)\n"
            f"  • 停车费: 匹配 {result['park_rows']} 条,未匹配 {result['park_unmatched']} 条\n"
            f"  • 违  章: 匹配 {result['vio_rows']} 条,未匹配 {result['vio_unmatched']} 条\n\n"
            f"输出文件:\n{self.dst_edit.text()}"
        )
        QMessageBox.information(self, "完成", msg)

    def on_failed(self, err: str):
        self.progress.setValue(0)
        self.status_label.setText("失败 ❌")
        self.match_btn.setEnabled(True)
        self.src_btn.setEnabled(True)
        self.dst_btn.setEnabled(True)
        QMessageBox.critical(self, "失败", f"匹配过程出错:\n\n{err}")


def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    win = MainWindow()
    win.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
